from django.contrib import admin
from core.admin_mixins import ImportExportAdminMixin
from django.utils.html import format_html
from django.urls import reverse, path
from django.utils.safestring import mark_safe
from django.contrib.admin import SimpleListFilter
from django.db.models import Count, Q
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from datetime import datetime
import io
import zipfile
from io import BytesIO
try:
    from openpyxl import load_workbook, Workbook
except Exception:  # Graceful fallback if openpyxl missing at runtime
    load_workbook = None
    Workbook = None
from .models import CertificateTemplate, Certificate, UserSignature


class TemplateStatusFilter(SimpleListFilter):
    title = 'حالة القالب'
    parameter_name = 'is_active'

    def lookups(self, request, model_admin):
        return (
            ('active', 'نشط'),
            ('inactive', 'غير نشط'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'active':
            return queryset.filter(is_active=True)
        elif self.value() == 'inactive':
            return queryset.filter(is_active=False)
        return queryset


class CertificateStatusFilter(SimpleListFilter):
    title = 'حالة الشهادة'
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return (
            ('active', 'نشطة'),
            ('revoked', 'ملغية'),
            ('expired', 'منتهية الصلاحية'),
        )

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset


class VerificationStatusFilter(SimpleListFilter):
    title = 'حالة التحقق'
    parameter_name = 'verification_status'

    def lookups(self, request, model_admin):
        return (
            ('verified', 'تم التحقق'),
            ('pending', 'في انتظار التحقق'),
            ('failed', 'فشل التحقق'),
        )

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(verification_status=self.value())
        return queryset


@admin.register(CertificateTemplate)
class CertificateTemplateAdmin(ImportExportAdminMixin, admin.ModelAdmin):
    list_display = (
        'template_name', 'institution_name', 'template_file_preview',
        'usage_count', 'default_status', 'is_active', 'created_at'
    )
    list_filter = (
        TemplateStatusFilter, 'is_default', 'is_active', 'created_at'
    )
    search_fields = ('template_name', 'institution_name', 'certificate_text')
    readonly_fields = ('created_at', 'updated_at', 'usage_count')
    
    fieldsets = (
        ('معلومات أساسية', {
            'fields': ('template_name', 'institution_name', 'institution_logo')
        }),
        ('القالب الجاهز', {
            'fields': ('template_file',)
        }),
        ('التوقيع', {
            'fields': (
                ('signature_name', 'signature_title'),
                'signature_image'
            )
        }),
        ('محتوى الشهادة', {
            'fields': ('certificate_text',)
        }),
        ('خيارات الإضافة', {
            'fields': (
                ('include_qr_code', 'include_grade'),
                ('include_completion_date', 'include_course_duration')
            )
        }),
        ('الحالة', {
            'fields': (
                ('is_default', 'is_active')
            )
        }),
        ('الإحصائيات', {
            'fields': ('usage_count',),
            'classes': ('collapse',)
        }),
        ('التواريخ', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def template_file_preview(self, obj):
        if obj.template_file:
            if obj.template_file.name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                return format_html(
                    '<img src="{}" width="50" height="30" style="border: 1px solid #ccc; object-fit: cover;" />',
                    obj.template_file.url
                )
            else:
                return format_html(
                    '<span style="color: #007bff;">📄 {}</span>',
                    obj.template_file.name.split('/')[-1]
                )
        return 'لا يوجد'
    template_file_preview.short_description = 'القالب'
    
    def usage_count(self, obj):
        count = obj.certificate_set.count()
        if count > 0:
            url = reverse('admin:certificates_certificate_changelist') + f'?template__id__exact={obj.id}'
            return format_html('<a href="{}">{} شهادة</a>', url, count)
        return '0 شهادة'
    usage_count.short_description = 'عدد الاستخدامات'
    
    def default_status(self, obj):
        if obj.is_default:
            return format_html('<span style="color: #28a745; font-weight: bold;">⭐ افتراضي</span>')
        return '⚪ عادي'
    default_status.short_description = 'افتراضي'
    
    
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.prefetch_related('certificate_set')
    
    actions = ['make_default', 'duplicate_template']
    
    def make_default(self, request, queryset):
        if queryset.count() > 1:
            self.message_user(request, 'يمكن تعيين قالب واحد فقط كافتراضي.', level='error')
            return
        
        template = queryset.first()
        CertificateTemplate.objects.filter(is_default=True).update(is_default=False)
        template.is_default = True
        template.save()
        self.message_user(request, f'تم تعيين "{template.template_name}" كقالب افتراضي.')
    make_default.short_description = "تعيين كقالب افتراضي"
    
    def duplicate_template(self, request, queryset):
        duplicated_count = 0
        for template in queryset:
            new_name = f"{template.template_name} - نسخة"
            # إنشاء نسخة بسيطة من القالب
            new_template = CertificateTemplate.objects.create(
                template_name=new_name,
                institution_name=template.institution_name,
                institution_logo=template.institution_logo,
                signature_name=template.signature_name,
                signature_title=template.signature_title,
                signature_image=template.signature_image,
                template_file=template.template_file,
                certificate_text=template.certificate_text,
                include_qr_code=template.include_qr_code,
                include_grade=template.include_grade,
                include_completion_date=template.include_completion_date,
                include_course_duration=template.include_course_duration,
                is_default=False,
                is_active=True
            )
            duplicated_count += 1
        
        self.message_user(request, f'تم إنشاء {duplicated_count} نسخة من القوالب.')
    duplicate_template.short_description = "إنشاء نسخة من القوالب"


@admin.register(Certificate)
class CertificateAdmin(admin.ModelAdmin):
    change_list_template = 'admin/certificates/certificate/change_list.html'
    list_display = (
        'certificate_id', 'student_name', 'national_id', 'course_title', 'final_grade_display',
        'status_display', 'verification_display', 'date_issued', 'preview_link', 'actions_column'
    )
    list_filter = (
        CertificateStatusFilter, VerificationStatusFilter, 'date_issued',
        'completion_date', 'course', 'template'
    )
    search_fields = (
        'certificate_id', 'student_name', 'course_title', 'user__username',
        'user__first_name', 'user__last_name', 'verification_code'
    )
    readonly_fields = (
        'certificate_id', 'verification_code', 'date_issued', 'created_at',
        'updated_at', 'qr_code_preview', 'verification_url_display'
    )
    
    fieldsets = (
        ('معلومات الشهادة', {
            'fields': (
                'certificate_id', 'user', 'course', 'template'
            )
        }),
        ('بيانات الطالب والدورة', {
            'fields': (
                'student_name', 'national_id', 'course_title', 'institution_name',
                'duration_days', 'course_duration_hours',
                ('start_date', 'end_date'), ('start_date_hijri', 'end_date_hijri'),
                'completion_date'
            )
        }),
        ('الدرجات والأداء', {
            'fields': (
                'final_grade', 'completion_percentage'
            )
        }),
        ('الحالة والتحقق', {
            'fields': (
                ('status', 'verification_status'),
                'verification_code', 'verification_url_display'
            )
        }),
        ('الملفات', {
            'fields': (
                'pdf_file', 'qr_code_image', 'qr_code_preview'
            )
        }),
        ('التوقيع الرقمي', {
            'fields': (
                'digital_signature', 'signature_verified', 'issued_by'
            ),
            'classes': ('collapse',)
        }),
        ('التواريخ', {
            'fields': ('date_issued', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    # --- Excel Import Utilities ---
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='certificates_certificate_import_excel'),
            path('import-excel/template/', self.admin_site.admin_view(self.download_excel_template), name='certificates_certificate_import_excel_template'),
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view), name='certificates_certificate_export_excel'),
        ]
        return custom_urls + urls

    def _parse_date(self, value):
        """Parse date cell (datetime or string) into datetime or return None."""
        if value is None or value == '':
            return None
        if isinstance(value, datetime):
            return value
        # Try common formats
        for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
        return None

    def download_excel_template(self, request):
        """Provide a simple Excel template with expected headers.
        Supports update by `certificate_id` or creation by `email` + (`course_id` or `course_title`).
        """
        headers = [
            # Update existing by certificate_id (preferred)
            'certificate_id',
            # Create or resolve user
            'email', 'student_name', 'national_id',
            # Resolve course
            'course_id', 'course_title',
            # Certificate content fields
            'duration_days', 'course_duration_hours', 'start_date', 'end_date',
            'start_date_hijri', 'end_date_hijri', 'completion_date',
            'final_grade', 'completion_percentage', 'status', 'verification_status',
            # Optional template resolution
            'template_id'
        ]

        if Workbook:
            wb = Workbook()
            ws = wb.active
            ws.title = 'الشهادات'
            # إضافة الرؤوس كـ strings صريحة
            ws.append([str(h) for h in headers])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8'
            )
            response['Content-Disposition'] = 'attachment; filename="certificates_import_template.xlsx"'
            response['Content-Encoding'] = 'utf-8'
            return response
        else:
            # Fallback CSV template
            csv_content = ','.join(headers) + '\n'
            return HttpResponse(csv_content, content_type='text/csv')

    def export_excel_view(self, request):
        """Export current certificates queryset to an Excel file with headers aligned to the import template."""
        headers = [
            'certificate_id', 'email', 'student_name', 'national_id',
            'course_id', 'course_title',
            'duration_days', 'course_duration_hours',
            'start_date', 'end_date', 'start_date_hijri', 'end_date_hijri', 'completion_date',
            'final_grade', 'completion_percentage', 'status', 'verification_status',
            'template_id'
        ]

        queryset = self.get_queryset(request)

        # دالة شاملة لإصلاح ترميز النصوص العربية (يجب تعريفها خارج if/else)
        def fix_arabic_encoding(text):
            """إصلاح ترميز النص العربي بطرق متعددة"""
            if not text:
                return ''
            
            # تحويل إلى string إذا لم يكن string
            if not isinstance(text, str):
                try:
                    text = str(text)
                except:
                    return ''
            
            # إذا كان النص يحتوي على رموز مثل Ø§Ø³Ø§، فهذا يعني أنه UTF-8 مخزن كـ Latin-1
            if 'Ø' in text or '§' in text or '³' in text or 'Ù' in text:
                try:
                    # الطريقة 1: تحويل من Latin-1 إلى bytes ثم decode كـ UTF-8
                    fixed = text.encode('latin-1').decode('utf-8')
                    # التحقق من أن الإصلاح نجح (لا يجب أن يحتوي على رموز غريبة)
                    if 'Ø' not in fixed and '§' not in fixed:
                        return fixed
                except (UnicodeEncodeError, UnicodeDecodeError):
                    pass
                
                try:
                    # الطريقة 2: استخدام errors='replace' أو 'ignore'
                    fixed = text.encode('latin-1', errors='ignore').decode('utf-8', errors='ignore')
                    if 'Ø' not in fixed and '§' not in fixed:
                        return fixed
                except:
                    pass
            
            # إذا كان النص bytes، decodeه كـ UTF-8
            if isinstance(text, bytes):
                try:
                    return text.decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        return text.decode('utf-8', errors='ignore')
                    except:
                        return text.decode('latin-1', errors='ignore')
            
            return text
        
        def safe_str(value):
            """تحويل القيمة إلى string بشكل آمن مع دعم العربية"""
            if value is None:
                return ''
            text = str(value) if not isinstance(value, str) else value
            return fix_arabic_encoding(text)

        if Workbook:
            wb = Workbook()
            ws = wb.active
            ws.title = 'الشهادات'
            
            # إضافة الرؤوس
            ws.append(headers)

            # إضافة البيانات مع التأكد من الترميز الصحيح للنصوص العربية
            from openpyxl.styles import Font
            
            # تعيين خط يدعم العربية للرؤوس
            header_font = Font(name='Arial', size=11, bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = str(header)
                cell.font = header_font
            
            for cert in queryset:
                row_data = [
                    safe_str(cert.certificate_id),
                    safe_str(cert.user.email if cert.user and cert.user.email else ''),
                    safe_str(cert.student_name),
                    safe_str(cert.national_id),
                    cert.course.id if cert.course else '',
                    safe_str(cert.course_title if cert.course_title else (cert.course.title if cert.course and hasattr(cert.course, 'title') else '')),
                    cert.duration_days if cert.duration_days is not None else '',
                    cert.course_duration_hours if cert.course_duration_hours is not None else '',
                    cert.start_date.isoformat() if cert.start_date else '',
                    cert.end_date.isoformat() if cert.end_date else '',
                    safe_str(cert.start_date_hijri),
                    safe_str(cert.end_date_hijri),
                    cert.completion_date.isoformat() if cert.completion_date else '',
                    cert.final_grade if cert.final_grade is not None else '',
                    cert.completion_percentage if cert.completion_percentage is not None else '',
                    safe_str(cert.status),
                    safe_str(cert.verification_status),
                    cert.template.id if cert.template else ''
                ]
                ws.append(row_data)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="certificates_export.xlsx"'
            # Excel يتعامل مع UTF-8 تلقائياً في ملفات .xlsx، لا حاجة لـ Content-Encoding
            return response
        else:
            # Fallback CSV export مع UTF-8 BOM
            import codecs
            # إضافة BOM للـ UTF-8
            csv_content = codecs.BOM_UTF8.decode('utf-8')
            csv_content += ','.join(headers) + '\n'
            for cert in queryset:
                row = [
                    safe_str(cert.certificate_id),
                    safe_str(cert.user.email if cert.user and cert.user.email else ''),
                    safe_str(cert.student_name),
                    safe_str(cert.national_id),
                    str(cert.course.id if cert.course else ''),
                    safe_str(cert.course_title if cert.course_title else (cert.course.title if cert.course and hasattr(cert.course, 'title') else '')),
                    str(cert.duration_days if cert.duration_days is not None else ''),
                    str(cert.course_duration_hours if cert.course_duration_hours is not None else ''),
                    str(cert.start_date.isoformat() if cert.start_date else ''),
                    str(cert.end_date.isoformat() if cert.end_date else ''),
                    safe_str(cert.start_date_hijri),
                    safe_str(cert.end_date_hijri),
                    str(cert.completion_date.isoformat() if cert.completion_date else ''),
                    str(cert.final_grade if cert.final_grade is not None else ''),
                    str(cert.completion_percentage if cert.completion_percentage is not None else ''),
                    safe_str(cert.status),
                    safe_str(cert.verification_status),
                    str(cert.template.id if cert.template else '')
                ]
                csv_content += ','.join(row) + '\n'
            return HttpResponse(csv_content.encode('utf-8'), content_type='text/csv; charset=utf-8')

    def import_excel_view(self, request):
        """Handle Excel upload to update or create certificate records.
        - If `certificate_id` exists: update that certificate.
        - Else if `email` exists: create/update for that user (optionally linking course by `course_id` or `course_title`).
        """
        base_ctx = self.admin_site.each_context(request)
        context = {
            **base_ctx,
            'title': 'استيراد بيانات الشهادات من ملف إكسل',
            'app_label': 'certificates',
            'model_name': 'certificate',
            'opts': self.model._meta,
            'template_url': reverse('admin:certificates_certificate_import_excel_template'),
            'changelist_url': reverse('admin:certificates_certificate_changelist'),
        }

        if request.method == 'POST':
            file = request.FILES.get('excel_file')
            if not file:
                messages.error(request, 'يرجى اختيار ملف إكسل.')
                return redirect('admin:certificates_certificate_import_excel')

            if not load_workbook:
                messages.error(request, 'حزمة openpyxl غير متاحة. يرجى تثبيتها أو استخدام قالب CSV.')
                return redirect('admin:certificates_certificate_import_excel')

            try:
                wb = load_workbook(filename=file, data_only=True)
                ws = wb.active
            except Exception as e:
                messages.error(request, f'فشل قراءة الملف: {e}')
                return redirect('admin:certificates_certificate_import_excel')

            # Map headers
            headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
            header_index = {h: i for i, h in enumerate(headers)}

            # Allow either update-by-certificate_id or create-by-email
            if 'certificate_id' not in header_index and 'email' not in header_index:
                messages.error(request, 'يجب أن يحتوي الصف الأول على "certificate_id" للتحديث أو "email" للإنشاء.')
                return redirect('admin:certificates_certificate_import_excel')

            updated_count = 0
            created_count = 0
            skipped_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                cert = None
                cert_id = row[header_index['certificate_id']] if 'certificate_id' in header_index else None
                # Try update by certificate_id
                if cert_id:
                    cert = Certificate.objects.filter(certificate_id=str(cert_id).strip()).first()
                
                # Else try create by email
                if not cert:
                    email = row[header_index['email']] if 'email' in header_index else None
                    if not email:
                        skipped_count += 1
                        continue

                    # Resolve or create user by email
                    from django.contrib.auth import get_user_model
                    UserModel = get_user_model()
                    user = UserModel.objects.filter(email=str(email).strip()).first()
                    if not user:
                        # Create a minimal user
                        username_base = str(email).split('@')[0]
                        username = username_base
                        suffix = 1
                        while UserModel.objects.filter(username=username).exists():
                            username = f"{username_base}{suffix}"
                            suffix += 1
                        user = UserModel.objects.create(
                            username=username,
                            email=str(email).strip(),
                            first_name=str(row[header_index['student_name']]).strip() if 'student_name' in header_index and row[header_index['student_name']] else '',
                            last_name=''  # يمكن تخصيص التقسيم لاحقاً إذا لزم
                        )

                    # Resolve course by id or title (optional)
                    course = None
                    try:
                        from courses.models import Course
                        if 'course_id' in header_index and row[header_index['course_id']]:
                            course_id_value = row[header_index['course_id']]
                            course = Course.objects.filter(id=int(str(course_id_value))).first()
                        elif 'course_title' in header_index and row[header_index['course_title']]:
                            course_title_value = str(row[header_index['course_title']]).strip()
                            course = Course.objects.filter(title__iexact=course_title_value).first()
                    except Exception:
                        course = None

                    # Resolve template if provided, else default
                    template = None
                    try:
                        from .models import CertificateTemplate
                        if 'template_id' in header_index and row[header_index['template_id']]:
                            template_id_value = row[header_index['template_id']]
                            template = CertificateTemplate.objects.filter(id=int(str(template_id_value))).first()
                        if not template:
                            template = CertificateTemplate.get_default_template() if hasattr(CertificateTemplate, 'get_default_template') else None
                    except Exception:
                        template = None

                    # Create certificate skeleton
                    cert = Certificate(
                        user=user,
                        course=course,
                        template=template,
                    )

                # Optional setters
                def set_field(name, transformer=lambda v: v):
                    if name in header_index:
                        value = row[header_index[name]]
                        if value is not None and value != '':
                            try:
                                setattr(cert, name, transformer(value))
                            except Exception:
                                pass

                set_field('student_name', str)
                set_field('national_id', str)
                set_field('course_title', str)
                set_field('duration_days', lambda v: int(str(v)))
                set_field('course_duration_hours', lambda v: int(str(v)))
                set_field('start_date', self._parse_date)
                set_field('end_date', self._parse_date)
                set_field('start_date_hijri', str)
                set_field('end_date_hijri', str)
                set_field('completion_date', self._parse_date)
                set_field('final_grade', lambda v: float(str(v)))
                set_field('completion_percentage', lambda v: float(str(v)))
                set_field('status', str)
                set_field('verification_status', str)

                try:
                    # Decide created vs updated
                    is_new = cert.pk is None
                    cert.save()
                    if is_new:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception:
                    skipped_count += 1

            messages.success(request, f'تم تحديث {updated_count} شهادة، وإنشاء {created_count} شهادة جديدة. تم تجاوز {skipped_count} صفوف.')
            return redirect('admin:certificates_certificate_changelist')

        return TemplateResponse(request, 'admin/import_export/upload.html', context)
    
    def final_grade_display(self, obj):
        if obj.final_grade is not None:
            if obj.final_grade >= 90:
                color = '#28a745'
                grade_text = 'ممتاز'
            elif obj.final_grade >= 80:
                color = '#007bff'
                grade_text = 'جيد جداً'
            elif obj.final_grade >= 70:
                color = '#ffc107'
                grade_text = 'جيد'
            elif obj.final_grade >= 60:
                color = '#fd7e14'
                grade_text = 'مقبول'
            else:
                color = '#dc3545'
                grade_text = 'ضعيف'
            
            grade_value = f"{obj.final_grade:.1f}%"
            return format_html(
                '<span style="color: {}; font-weight: bold;">{} ({})</span>',
                color, grade_value, grade_text
            )
        return 'غير محدد'
    final_grade_display.short_description = 'الدرجة النهائية'
    
    def status_display(self, obj):
        status_colors = {
            'active': ('#28a745', '✅ نشطة'),
            'revoked': ('#dc3545', '❌ ملغية'),
            'expired': ('#6c757d', '⏰ منتهية'),
        }
        color, text = status_colors.get(obj.status, ('#6c757d', obj.get_status_display()))
        return format_html('<span style="color: {}; font-weight: bold;">{}</span>', color, text)
    status_display.short_description = 'الحالة'
    
    def verification_display(self, obj):
        verification_colors = {
            'verified': ('#28a745', '✅ تم التحقق'),
            'pending': ('#ffc107', '⏳ في الانتظار'),
            'failed': ('#dc3545', '❌ فشل'),
        }
        color, text = verification_colors.get(obj.verification_status, ('#6c757d', obj.get_verification_status_display()))
        return format_html('<span style="color: {}; font-weight: bold;">{}</span>', color, text)
    verification_display.short_description = 'التحقق'
    
    def qr_code_preview(self, obj):
        if obj.qr_code_image:
            return format_html(
                '<img src="{}" width="100" height="100" style="border: 1px solid #ccc;" />',
                obj.qr_code_image.url
            )
        return 'لا يوجد'
    qr_code_preview.short_description = 'معاينة رمز QR'
    
    def verification_url_display(self, obj):
        url = obj.get_verification_url()
        return format_html('<a href="{}" target="_blank">{}</a>', url, url)
    verification_url_display.short_description = 'رابط التحقق'
    
    def preview_link(self, obj):
        """عرض رابط المعاينة للشهادة"""
        if obj.verification_code:
            url = obj.get_verification_url()
            return format_html(
                '<a href="{}" target="_blank" style="color: #0e5181; font-weight: 600; text-decoration: none; padding: 4px 8px; background: #e3f2fd; border-radius: 4px; display: inline-block;">👁️ معاينة</a>',
                url
            )
        return format_html('<span style="color: #999;">لا يوجد</span>')
    preview_link.short_description = 'معاينة'
    preview_link.admin_order_field = 'verification_code'
    
    def actions_column(self, obj):
        actions = []
        
        if obj.status == 'active':
            actions.append(
                f'<a href="#" onclick="revokeCertificate({obj.id})" style="color: #dc3545;">إلغاء</a>'
            )
        
        # رابط معاينة PDF بدلاً من صفحة التحقق العادية
        if obj.verification_code:
            pdf_preview_url = reverse('certificate_pdf_preview', args=[obj.verification_code])
            actions.append(
                format_html(
                    '<a href="{}" target="_blank" style="color: #28a745;">تحقق</a>',
                    pdf_preview_url
                )
            )
        
        return format_html(' | '.join(actions))
    actions_column.short_description = 'الإجراءات'
    
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.select_related('user', 'course', 'template', 'issued_by')
    
    actions = ['delete_selected', 'download_certificates_images']
    
    
    def download_certificates_images(self, request, queryset):
        """تحميل الشهادات المحددة كصور PNG مضغوطة في ZIP"""
        # التحقق من وجود شهادات محددة
        if queryset.count() == 0:
            self.message_user(request, 'لم يتم تحديد أي شهادات للتحميل.', level='warning')
            return
        
        try:
            from .utils import generate_certificate_image_from_url, PLAYWRIGHT_AVAILABLE
            
            if not PLAYWRIGHT_AVAILABLE:
                self.message_user(
                    request,
                    'خطأ: مكتبة Playwright غير مثبتة. يرجى تثبيتها باستخدام: pip install playwright && playwright install chromium',
                    level='error'
                )
                return
            
            # إنشاء ملف ZIP في الذاكرة
            zip_buffer = BytesIO()
            errors = []
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                count = 0
                total = queryset.count()
                
                for index, certificate in enumerate(queryset, 1):
                    try:
                        if not certificate.verification_code:
                            errors.append(f"الشهادة {certificate.certificate_id}: لا يوجد رمز تحقق")
                            continue
                        
                        # الحصول على رابط التحقق
                        verify_url = certificate.get_verification_url()
                        
                        # تحويل صفحة التحقق إلى صورة
                        image_data = generate_certificate_image_from_url(verify_url)
                        
                        if image_data:
                            # إضافة الصورة إلى ZIP
                            safe_name = certificate.student_name.replace('/', '_').replace('\\', '_') if certificate.student_name else 'unknown'
                            filename = f"certificate_{certificate.certificate_id}_{safe_name}.png"
                            # تنظيف اسم الملف من الأحرف غير المسموحة
                            filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).strip()
                            zip_file.writestr(filename, image_data)
                            count += 1
                        else:
                            errors.append(f"الشهادة {certificate.certificate_id}: فشل إنشاء الصورة")
                    except Exception as e:
                        # تخطي الشهادات التي فشل إنشاء الصورة لها
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error generating image for certificate {certificate.certificate_id}: {str(e)}")
                        errors.append(f"الشهادة {certificate.certificate_id}: {str(e)}")
                        continue
            
            if count == 0:
                error_msg = 'لم يتم إنشاء أي صورة.'
                if errors:
                    error_msg += f' الأخطاء: {"; ".join(errors[:5])}'
                self.message_user(
                    request,
                    error_msg,
                    level='warning'
                )
                return
            
            # إرجاع ملف ZIP
            zip_buffer.seek(0)
            response = HttpResponse(
                zip_buffer.read(),
                content_type='application/zip'
            )
            response['Content-Disposition'] = f'attachment; filename="certificates_images_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip"'
            
            success_msg = f'تم تحميل {count} صورة شهادة بنجاح.'
            if errors:
                success_msg += f' (فشل {len(errors)} شهادة)'
            self.message_user(request, success_msg)
            return response
            
        except ImportError as e:
            self.message_user(
                request,
                f'خطأ: {str(e)}',
                level='error'
            )
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception("Error in download_certificates_images")
            self.message_user(
                request,
                f'حدث خطأ أثناء إنشاء ملف ZIP: {str(e)}',
                level='error'
            )
    download_certificates_images.short_description = "تحميل الشهادات المحددة كصور"


@admin.register(UserSignature)
class UserSignatureAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'signature_name', 'signature_title', 'signature_preview',
        'default_status', 'is_active', 'created_at'
    )
    list_filter = ('is_default', 'is_active', 'created_at')
    search_fields = ('user__username', 'signature_name', 'signature_title')
    readonly_fields = ('created_at', 'signature_preview')
    
    fieldsets = (
        ('معلومات التوقيع', {
            'fields': ('user', 'signature_name', 'signature_title')
        }),
        ('صورة التوقيع', {
            'fields': ('signature_image', 'signature_preview')
        }),
        ('الحالة', {
            'fields': (('is_default', 'is_active'),)
        }),
        ('التواريخ', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )
    
    def signature_preview(self, obj):
        if obj.signature_image:
            return format_html(
                '<img src="{}" width="150" height="75" style="border: 1px solid #ccc; object-fit: contain;" />',
                obj.signature_image.url
            )
        return 'لا يوجد'
    signature_preview.short_description = 'معاينة التوقيع'
    
    def default_status(self, obj):
        if obj.is_default:
            return format_html('<span style="color: #28a745; font-weight: bold;">⭐ افتراضي</span>')
        return '⚪ عادي'
    default_status.short_description = 'افتراضي'
    
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.select_related('user')


