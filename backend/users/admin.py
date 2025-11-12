from django.contrib import admin
from core.admin_mixins import ImportExportAdminMixin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import User
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.contrib.admin import SimpleListFilter
from django.db.models import Count, Q
from django.conf import settings
from .models import Profile, Organization, Instructor, Student, ArchivedUser
from .forms import CustomUserCreationForm, CustomUserChangeForm
from django.utils import timezone
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import path
from django.contrib import messages
from django.shortcuts import redirect
from django.utils.translation import gettext_lazy as _
import io
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

class StatusFilter(SimpleListFilter):
    title = 'حالة المستخدم'
    parameter_name = 'profile__status'

    def lookups(self, request, model_admin):
        return (
            ('Student', 'طالب'),
            ('Instructor', 'مدرب'),
            ('Admin', 'مدير'),
            ('Organization', 'منظمة'),
        )

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(profile__status=self.value())
        return queryset


class ProfileInline(admin.StackedInline):
    model = Profile
    can_delete = False
    verbose_name_plural = 'الملف الشخصي'
    fields = (
        ('name', 'status'),
        'national_id',
        ('email', 'phone'),
        'image_profile',
        'shortBio',
        'detail',
        ('github', 'youtube'),
        ('twitter', 'facebook'),
        ('instagram', 'linkedin'),
        'is_archived',
    )
    readonly_fields = ('is_archived',)


class InstructorInline(admin.StackedInline):
    model = Instructor
    can_delete = False
    verbose_name_plural = 'بيانات المدرب'
    fields = (
        ('organization', 'department'),
        ('qualification', 'date_of_birth'),
        'bio',
        'research_interests'
    )
    
    def has_add_permission(self, request, obj=None):
        # Only show for instructors
        if obj and hasattr(obj, 'profile') and obj.profile.status == 'Instructor':
            return True
        return False


class StudentInline(admin.StackedInline):
    model = Student
    can_delete = False
    verbose_name_plural = 'بيانات الطالب'
    fields = ('department', 'date_of_birth')
    
    def has_add_permission(self, request, obj=None):
        # Only show for students
        if obj and hasattr(obj, 'profile') and obj.profile.status == 'Student':
            return True
        return False


# Unregister the default User admin
admin.site.unregister(User)


@admin.register(User)
class CustomUserAdmin(BaseUserAdmin):
    change_list_template = 'admin/auth/user/change_list.html'
    add_form = CustomUserCreationForm
    form = CustomUserChangeForm
    inlines = (InstructorInline, StudentInline) if hasattr(settings, 'SHOW_ALL_INLINES') else ()
    actions = ['archive_selected_users']
    list_display = (
        'username', 'email', 'national_id_display', 'first_name', 'last_name', 
        'status_dropdown', 'profile_image', 'is_active', 
        'courses_count', 'date_joined', 'archive_button'
    )
    list_filter = (
        StatusFilter, 'is_active', 'is_staff', 'is_superuser', 'date_joined'
    )
    search_fields = ('username', 'first_name', 'last_name', 'email', 'profile__name', 'profile__national_id')
    
    # إزالة قسم الصلاحيات من fieldsets
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("المعلومات الشخصية", {"fields": ("first_name", "last_name", "email", "national_id", "phone")}),
        ("التواريخ المهمة", {"fields": ("last_login", "date_joined")}),
    )
    
    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": ("email", "first_name", "last_name", "national_id", "phone", "username", "password1", "password2"),
            },
        ),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='auth_user_import_excel'),
            path('import-excel/template/', self.admin_site.admin_view(self.download_excel_template), name='auth_user_import_excel_template'),
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view), name='auth_user_export_excel'),
            path('<int:user_id>/update-status/', self.admin_site.admin_view(self.update_user_status), name='auth_user_update_status'),
            path('<int:user_id>/archive/', self.admin_site.admin_view(self.archive_user_view), name='auth_user_archive'),
        ]
        return custom_urls + urls

    def download_excel_template(self, request):
        headers = [
            'id', 'username', 'email', 'first_name', 'last_name',
            'is_active', 'is_staff', 'is_superuser', 'password'
        ]
        if Workbook:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Users'
            ws.append(headers)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="users_import_template.xlsx"'
            return response
        else:
            csv_content = ','.join(headers) + '\n'
            return HttpResponse(csv_content, content_type='text/csv')

    def import_excel_view(self, request):
        base_ctx = self.admin_site.each_context(request)
        context = {
            **base_ctx,
            'title': 'استيراد المستخدمين من ملف إكسل',
            'app_label': 'auth',
            'model_name': 'user',
            'opts': self.model._meta,
            'template_url': reverse('admin:auth_user_import_excel_template'),
            'changelist_url': reverse('admin:auth_user_changelist'),
        }

        if request.method == 'POST':
            file = request.FILES.get('excel_file')
            if not file:
                messages.error(request, 'يرجى اختيار ملف إكسل.')
                return redirect('admin:auth_user_import_excel')
            if not load_workbook:
                messages.error(request, 'حزمة openpyxl غير متاحة. يرجى تثبيتها أو استخدام قالب CSV.')
                return redirect('admin:auth_user_import_excel')

            try:
                wb = load_workbook(filename=file, data_only=True)
                ws = wb.active
            except Exception as e:
                messages.error(request, f'فشل قراءة الملف: {e}')
                return redirect('admin:auth_user_import_excel')

            # Map headers
            header_row = None
            for row in ws.iter_rows(values_only=True):
                header_row = row
                break
            if not header_row:
                messages.error(request, 'الملف لا يحتوي على صف رؤوس.')
                return redirect('admin:auth_user_import_excel')

            header_index = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}
            required = ['email']
            if not any(col in header_index for col in ['id', 'email', 'username']):
                messages.error(request, 'يجب أن يحتوي الملف على عمود واحد على الأقل للتعرف على المستخدم: id أو email أو username.')
                return redirect('admin:auth_user_import_excel')

            created, updated = 0, 0
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx == 0:
                    continue
                if not any(row):
                    continue

                user_obj = None
                # Resolve by id
                if 'id' in header_index and row[header_index['id']]:
                    try:
                        user_obj = User.objects.filter(id=int(str(row[header_index['id']]))).first()
                    except Exception:
                        user_obj = None

                # Else by email
                if not user_obj and 'email' in header_index and row[header_index['email']]:
                    email_val = str(row[header_index['email']]).strip().lower()
                    user_obj = User.objects.filter(email__iexact=email_val).first()

                # Else by username
                if not user_obj and 'username' in header_index and row[header_index['username']]:
                    username_val = str(row[header_index['username']]).strip()
                    user_obj = User.objects.filter(username__iexact=username_val).first()

                # Create if not exists
                if not user_obj:
                    username_val = str(row[header_index['username']]).strip() if 'username' in header_index and row[header_index['username']] else None
                    email_val = str(row[header_index['email']]).strip().lower() if 'email' in header_index and row[header_index['email']] else None
                    if not username_val:
                        username_val = (email_val.split('@')[0] if email_val else f'user_{timezone.now().strftime("%Y%m%d%H%M%S")}')
                    user_obj = User(username=username_val, email=email_val)
                    created += 1
                else:
                    updated += 1

                # Set attributes
                if 'first_name' in header_index and row[header_index['first_name']]:
                    user_obj.first_name = str(row[header_index['first_name']]).strip()
                if 'last_name' in header_index and row[header_index['last_name']]:
                    user_obj.last_name = str(row[header_index['last_name']]).strip()
                if 'email' in header_index and row[header_index['email']]:
                    user_obj.email = str(row[header_index['email']]).strip().lower()
                if 'is_active' in header_index and row[header_index['is_active']] is not None:
                    user_obj.is_active = str(row[header_index['is_active']]).strip().lower() in ['1', 'true', 'yes']
                if 'is_staff' in header_index and row[header_index['is_staff']] is not None:
                    user_obj.is_staff = str(row[header_index['is_staff']]).strip().lower() in ['1', 'true', 'yes']
                if 'is_superuser' in header_index and row[header_index['is_superuser']] is not None:
                    user_obj.is_superuser = str(row[header_index['is_superuser']]).strip().lower() in ['1', 'true', 'yes']

                # Set password if provided
                if 'password' in header_index and row[header_index['password']]:
                    user_obj.set_password(str(row[header_index['password']]))

                user_obj.save()

            messages.success(request, f'تم الاستيراد بنجاح: تم إنشاء {created} مستخدم وتحديث {updated} مستخدم.')
            return redirect('admin:auth_user_changelist')

        return TemplateResponse(request, 'admin/import_export/upload.html', context)

    def export_excel_view(self, request):
        headers = [
            'id', 'username', 'email', 'first_name', 'last_name',
            'is_active', 'is_staff', 'is_superuser', 'date_joined', 'last_login'
        ]
        queryset = self.get_queryset(request)

        if Workbook:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Users'
            ws.append(headers)
            for u in queryset:
                ws.append([
                    u.id, u.username, u.email or '', u.first_name or '', u.last_name or '',
                    u.is_active, u.is_staff, u.is_superuser,
                    u.date_joined.isoformat() if u.date_joined else '',
                    u.last_login.isoformat() if u.last_login else ''
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
            return response
        else:
            rows = []
            rows.append(','.join(headers))
            for u in queryset:
                row = [
                    str(u.id), u.username,
                    str(u.email or ''), str(u.first_name or ''), str(u.last_name or ''),
                    '1' if u.is_active else '0', '1' if u.is_staff else '0', '1' if u.is_superuser else '0',
                    str(u.date_joined.isoformat() if u.date_joined else ''),
                    str(u.last_login.isoformat() if u.last_login else '')
                ]
                rows.append(','.join(row))
            csv_content = '\n'.join(rows)
            return HttpResponse(csv_content, content_type='text/csv')
    
    def status_dropdown(self, obj):
        """عرض dropdown قابل للتعديل للحالة"""
        try:
            profile = obj.profile
            current_status = profile.status or 'Student'
        except Profile.DoesNotExist:
            current_status = 'Student'
        
        status_choices = [
            ('Student', 'طالب'),
            ('Instructor', 'مدرب'),
            ('Admin', 'مدير'),
         
        ]
        
        status_colors = {
            'Student': '#28a745',
            'Instructor': '#007bff',
            'Admin': '#dc3545',
           
        }
        
        color = status_colors.get(current_status, '#6c757d')
        
        # بناء options HTML بشكل مباشر
        options_html = ''
        for value, label in status_choices:
            selected = ' selected' if value == current_status else ''
            options_html += f'<option value="{value}"{selected}>{label}</option>'
        
        update_url = reverse('admin:auth_user_update_status', args=[obj.id])
        
        # بناء HTML كامل باستخدام mark_safe
        html = (
            f'<select class="status-dropdown" data-user-id="{obj.id}" data-url="{update_url}" '
            f'style="padding: 4px 8px; border: 1px solid #ddd; border-radius: 4px; '
            f'background-color: white; color: {color}; font-weight: bold; cursor: pointer; '
            f'min-width: 100px;" onchange="updateUserStatus(this)">'
            f'{options_html}'
            f'</select>'
        )
        
        return mark_safe(html)
    status_dropdown.short_description = 'الحالة'
    
    def update_user_status(self, request, user_id):
        """تحديث حالة المستخدم"""
        from django.http import JsonResponse
        
        if request.method != 'POST':
            return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
        try:
            user = User.objects.get(id=user_id)
            new_status = request.POST.get('status')
            
            if new_status not in ['Student', 'Instructor', 'Admin', 'Organization']:
                return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
            
            profile, created = Profile.objects.get_or_create(user=user)
            profile.status = new_status
            profile.save()
            
            # الحصول على النص المعروض للحالة
            status_display = dict(Profile.status_choices).get(new_status, new_status)
            
            return JsonResponse({
                'success': True,
                'status': new_status,
                'status_display': status_display
            })
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    def profile_image(self, obj):
        try:
            if obj.profile.image_profile:
                return format_html(
                    '<img src="{}" width="30" height="30" style="border-radius: 50%;" />',
                    obj.profile.image_profile.url
                )
            return '📷'
        except (Profile.DoesNotExist, ValueError):
            return '❌'
    profile_image.short_description = 'الصورة'

    def national_id_display(self, obj):
        try:
            return obj.profile.national_id or '-'
        except Profile.DoesNotExist:
            return '-'
    national_id_display.short_description = 'رقم الهوية'
    
    def archive_button(self, obj):
        try:
            if obj.profile.is_archived:
                return format_html('<span style="color:#6c757d;font-weight:bold;">مؤرشف</span>')
        except Profile.DoesNotExist:
            pass

        archive_url = reverse('admin:auth_user_archive', args=[obj.pk])
        return format_html(
            '<button type="button" class="button archive-trigger" '
            'style="background:transparent;border:none;color:#dc3545;cursor:pointer;'
            'display:inline-flex;align-items:center;gap:6px;padding:4px 6px;" '
            'data-archive-url="{}" data-username="{}" '
            'title="مسح">'
            '🗑️</button>',
            archive_url,
            obj.get_full_name() or obj.username
        )
    archive_button.short_description = 'مسح'

    def courses_count(self, obj):
        try:
            if obj.profile.status == 'Instructor':
                count = obj.instructor.course_set.count() if hasattr(obj, 'instructor') else 0
                if count > 0:
                    url = reverse('admin:courses_course_changelist') + f'?instructor__profile__user__id__exact={obj.id}'
                    return format_html('<a href="{}">{} دورة</a>', url, count)
                return '0 دورة'
            elif obj.profile.status == 'Student':
                count = obj.course_enrollments.count()
                if count > 0:
                    url = reverse('admin:courses_enrollment_changelist') + f'?student__id__exact={obj.id}'
                    return format_html('<a href="{}">{} دورة</a>', url, count)
                return '0 دورة'
            return '-'
        except Profile.DoesNotExist:
            return '-'
    courses_count.short_description = 'الدورات'
    
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        queryset = queryset.select_related('profile').prefetch_related('course_enrollments')
        return queryset.filter(
            Q(profile__isnull=True) | Q(profile__is_archived=False)
        )

    def delete_model(self, request, obj):
        archived = self._archive_user(request, obj)
        if archived:
            self.message_user(request, f"تم أرشفة المستخدم {obj.get_full_name() or obj.username} بنجاح.", messages.SUCCESS)
        else:
            self.message_user(request, f"تعذرت أرشفة المستخدم {obj.get_full_name() or obj.username}.", messages.WARNING)

    def delete_queryset(self, request, queryset):
        archived_count = 0
        for user in queryset:
            if self._archive_user(request, user):
                archived_count += 1
        if archived_count:
            self.message_user(request, f"تم أرشفة {archived_count} مستخدم.", messages.SUCCESS)
        else:
            self.message_user(request, "لم يتم أرشفة أي مستخدم.", messages.WARNING)

    def archive_selected_users(self, request, queryset):
        archived_count = 0
        for user in queryset:
            if self._archive_user(request, user):
                archived_count += 1
        if archived_count:
            self.message_user(request, f"تم أرشفة {archived_count} مستخدم.", messages.SUCCESS)
        else:
            self.message_user(request, "لم يتم أرشفة أي مستخدم.", messages.WARNING)
    archive_selected_users.short_description = "أرشفة المستخدمين المحددين"

    def _archive_user(self, request, user_obj):
        if not user_obj.pk:
            return False

        profile = getattr(user_obj, 'profile', None)
        national_id = None
        profile_payload = None

        if profile:
            if profile.is_archived:
                return False
            national_id = profile.national_id
            profile_payload = {
                'name': profile.name,
                'status': profile.status,
                'email': profile.email,
                'phone': profile.phone,
                'image_profile': profile.image_profile.url if profile.image_profile else None,
                'shortBio': profile.shortBio,
                'detail': profile.detail,
                'social_links': {
                    'github': profile.github,
                    'youtube': profile.youtube,
                    'twitter': profile.twitter,
                    'facebook': profile.facebook,
                    'instagram': profile.instagram,
                    'linkedin': profile.linkedin,
                },
            }
            profile.is_archived = True
            profile.save(update_fields=['is_archived'])

        user_obj.is_active = False
        user_obj.save(update_fields=['is_active'])

        archived_by = request.user if hasattr(request, 'user') and request.user.is_authenticated else None

        ArchivedUser.objects.update_or_create(
            original_user=user_obj,
            defaults={
                'username': user_obj.username,
                'email': user_obj.email,
                'first_name': user_obj.first_name,
                'last_name': user_obj.last_name,
                'national_id': national_id,
                'profile_data': profile_payload,
                'archived_by': archived_by,
            }
        )
        return True

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        save_profile = getattr(form, "save_profile", None)
        if callable(save_profile):
            save_profile(obj)

    def response_add(self, request, obj, post_url_continue=None):
        """بعد حفظ المستخدم الجديد، العودة لقائمة المستخدمين مع رسالة نجاح."""
        self.message_user(request, _("تم حفظ المستخدم بنجاح."), messages.SUCCESS)
        return HttpResponseRedirect(reverse('admin:auth_user_changelist'))

    def response_change(self, request, obj):
        """بعد تعديل المستخدم، العودة لقائمة المستخدمين مع رسالة نجاح."""
        self.message_user(request, _("تم تحديث المستخدم بنجاح."), messages.SUCCESS)
        return HttpResponseRedirect(reverse('admin:auth_user_changelist'))

    def archive_user_view(self, request, user_id):
        if not self.has_delete_permission(request):
            self.message_user(request, "ليست لديك صلاحية لأرشفة المستخدمين.", level=messages.ERROR)
            return HttpResponseRedirect(reverse('admin:auth_user_changelist'))

        user = self.get_object(request, user_id)
        if not user:
            self.message_user(request, "المستخدم غير موجود.", level=messages.ERROR)
            return HttpResponseRedirect(reverse('admin:auth_user_changelist'))

        archived = self._archive_user(request, user)
        if archived:
            self.message_user(request, f"تم أرشفة المستخدم {user.get_full_name() or user.username}.", level=messages.SUCCESS)
        else:
            self.message_user(request, "المستخدم مؤرشف بالفعل أو تعذر الأرشفة.", level=messages.WARNING)
        return HttpResponseRedirect(reverse('admin:auth_user_changelist'))


@admin.register(Profile)
class ProfileAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'user_username', 'status', 'national_id', 'email', 'phone', 
        'profile_image', 'social_links', 'created_date'
    )
    list_filter = ('status', 'user__date_joined', 'is_archived')
    search_fields = ('name', 'user__username', 'email', 'phone', 'shortBio', 'national_id')
    readonly_fields = ('id', 'created_date', 'is_archived')
    
    fieldsets = (
        ('معلومات أساسية', {
            'fields': ('user', 'name', 'status', 'national_id', 'email', 'phone', 'is_archived')
        }),
        ('الصورة والوصف', {
            'fields': ('image_profile', 'shortBio', 'detail')
        }),
        ('الروابط الاجتماعية', {
            'fields': (
                ('github', 'youtube'),
                ('twitter', 'facebook'),
                ('instagram', 'linkedin')
            ),
            'classes': ('collapse',)
        }),
        ('معلومات النظام', {
            'fields': ('id', 'created_date'),
            'classes': ('collapse',)
        }),
    )
    
    def user_username(self, obj):
        if obj.user:
            url = reverse('admin:auth_user_change', args=[obj.user.id])
            return format_html('<a href="{}">{}</a>', url, obj.user.username)
        return '-'
    user_username.short_description = 'اسم المستخدم'
    
    def profile_image(self, obj):
        if obj.image_profile:
            return format_html(
                '<img src="{}" width="40" height="40" style="border-radius: 50%;" />',
                obj.image_profile.url
            )
        return '📷'
    profile_image.short_description = 'الصورة'
    
    def social_links(self, obj):
        links = []
        social_fields = {
            'github': '🐙',
            'youtube': '📺',
            'twitter': '🐦',
            'facebook': '📘',
            'instagram': '📷',
            'linkedin': '💼'
        }
        
        for field, emoji in social_fields.items():
            value = getattr(obj, field)
            if value:
                links.append(f'<a href="{value}" target="_blank">{emoji}</a>')
        
        return format_html(' '.join(links)) if links else '-'
    social_links.short_description = 'الروابط الاجتماعية'
    
    def created_date(self, obj):
        return obj.user.date_joined if obj.user else None
    created_date.short_description = 'تاريخ الإنشاء'
    
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.select_related('user')


@admin.register(Organization)
class OrganizationAdmin(admin.ModelAdmin):
    list_display = ('profile_name', 'location', 'website', 'founded_year', 'employees', 'teachers_count')
    list_filter = ('founded_year',)
    search_fields = ('profile__name', 'location', 'website')
    
    fieldsets = (
        ('معلومات أساسية', {
            'fields': ('profile', 'location', 'website', 'founded_year', 'employees')
        }),
        ('الوصف', {
            'fields': ('description',)
        }),
    )
    
    def profile_name(self, obj):
        if obj.profile:
            url = reverse('admin:users_profile_change', args=[obj.profile.id])
            return format_html('<a href="{}">{}</a>', url, obj.profile.name)
        return '-'
    profile_name.short_description = 'اسم المنظمة'
    
    def teachers_count(self, obj):
        count = obj.instructor_set.count()
        if count > 0:
            url = reverse('admin:users_instructor_changelist') + f'?organization__id__exact={obj.id}'
            return format_html('<a href="{}">{} معلم</a>', url, count)
        return '0 معلم'
    teachers_count.short_description = 'المعلمين'
    
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.select_related('profile').prefetch_related('instructor_set')

@admin.register(Instructor)
class InstructorAdmin(ImportExportAdminMixin, admin.ModelAdmin):
    list_display = (
        'profile_name', 'phone', 'email',
        'courses_count', 'students_count'
    )
    list_filter = ('organization', 'department', 'date_of_birth')
    search_fields = (
        'profile__name', 'profile__user__username',
        'profile__email', 'profile__phone', 'department', 'qualification'
    )

    fieldsets = (
        ('معلومات أساسية', {
            'fields': ('profile', 'organization', 'department', 'qualification', 'date_of_birth')
        }),
        ('السيرة الذاتية', {
            'fields': ('bio', 'research_interests')
        }),
    )

    def profile_name(self, obj):
        if obj.profile:
            url = reverse('admin:users_profile_change', args=[obj.profile.id])
            return format_html('<a href="{}">{}</a>', url, obj.profile.name)
        return '-'
    profile_name.short_description = 'اسم المدرب'

    def phone(self, obj):
        return getattr(obj.profile, 'phone', '-') if obj.profile else '-'
    phone.short_description = 'رقم الهاتف'

    def email(self, obj):
        return obj.profile.email if obj.profile and obj.profile.email else '-'
    email.short_description = 'البريد الإلكتروني'

    def courses_count(self, obj):
        count = obj.courses_taught.count()
        if count > 0:
            url = reverse('admin:courses_course_changelist') + f'?instructor__id__exact={obj.id}'
            return format_html('<a href="{}">{} دورة</a>', url, count)
        return '0 دورة'
    courses_count.short_description = 'الدورات'

    def students_count(self, obj):
        from django.db.models import Count
        count = obj.courses_taught.aggregate(total=Count('enrollments'))['total'] or 0
        if count > 0:
            url = reverse('admin:courses_enrollment_changelist') + f'?course__instructor__id__exact={obj.id}'
            return format_html('<a href="{}">{} طالب</a>', url, count)
        return '0 طالب'
    students_count.short_description = 'الطلاب'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('profile', 'organization').prefetch_related('courses_taught', 'courses_taught__enrollments')


@admin.register(Student)
class StudentAdmin(ImportExportAdminMixin, admin.ModelAdmin):
    list_display = ('profile_name', 'phone', 'email', 'enrolled_courses', 'completed_courses')
    list_filter = ('department', 'date_of_birth')
    search_fields = ('profile__name', 'profile__user__username', 'profile__email', 'profile__phone', 'department')

    fieldsets = (
        ('معلومات أساسية', {
            'fields': ('profile', 'department', 'date_of_birth')
        }),
    )

    def profile_name(self, obj):
        if obj.profile:
            url = reverse('admin:users_profile_change', args=[obj.profile.id])
            return format_html('<a href="{}">{}</a>', url, obj.profile.name)
        return '-'
    profile_name.short_description = 'اسم الطالب'

    def phone(self, obj):
        return getattr(obj.profile, 'phone', '-') if obj.profile else '-'
    phone.short_description = 'رقم الهاتف'

    def email(self, obj):
        return obj.profile.email if obj.profile and obj.profile.email else '-'
    email.short_description = 'البريد الإلكتروني'

    def enrolled_courses(self, obj):
        if obj.profile and obj.profile.user:
            count = obj.profile.user.course_enrollments.filter(status='active').count()
            if count > 0:
                url = reverse('admin:courses_enrollment_changelist') + f'?student__id__exact={obj.profile.user.id}'
                return format_html('<a href="{}">{} دورة</a>', url, count)
            return '0 دورة'
        return '-'
    enrolled_courses.short_description = 'الدورات المسجلة'

    def completed_courses(self, obj):
        if obj.profile and obj.profile.user:
            count = obj.profile.user.course_enrollments.filter(status='completed').count()
            return f'{count} دورة'
        return '0 دورة'
    completed_courses.short_description = 'الدورات المكتملة'

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.select_related('profile__user').prefetch_related('profile__user__course_enrollments')


@admin.register(ArchivedUser)
class ArchivedUserAdmin(admin.ModelAdmin):
    list_display = ('username', 'email', 'archived_at', 'archived_by')
    search_fields = ('username', 'email', 'first_name', 'last_name', 'national_id')
    readonly_fields = ('original_user', 'username', 'email', 'first_name', 'last_name', 'national_id', 'profile_data', 'archived_at', 'archived_by')

    def has_module_permission(self, request):
        return False

    def has_view_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
