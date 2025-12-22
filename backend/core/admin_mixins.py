from django.contrib import admin
from django.urls import path, reverse
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.shortcuts import redirect
from django.contrib import messages
from django.db import models
import io
from datetime import datetime
import uuid
from django.db.models.fields.files import FieldFile

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None


class ImportExportAdminMixin:
    """Generic mixin to add import/export (Excel/CSV fallback) to any ModelAdmin.

    - Adds routes: import-excel/, import-excel/template/, export-excel/
    - Uses `admin/import_export/upload.html` for uploads.
    - Injects context into change list to show buttons via custom template.
    - Operates on basic model fields; ForeignKey exported as `<field>_id`.
    """

    change_list_template = 'admin/change_list.html'

    def get_urls(self):
        opts = self.model._meta
        base = f"{opts.app_label}_{opts.model_name}"
        custom_urls = [
            path(
                'import-excel/',
                self.admin_site.admin_view(self.import_excel_view),
                name=f'{base}_import_excel',
            ),
            path(
                'import-excel/template/',
                self.admin_site.admin_view(self.download_excel_template),
                name=f'{base}_import_excel_template',
            ),
            path(
                'export-excel/',
                self.admin_site.admin_view(self.export_excel_view),
                name=f'{base}_export_excel',
            ),
        ]
        return custom_urls + super().get_urls()

    # ----- Helpers -----
    def _model_headers(self):
        fields = []
        for f in self.model._meta.get_fields():
            if getattr(f, 'auto_created', False):
                continue
            if f.many_to_many or f.one_to_many:
                continue
            # Store foreign keys as <name>_id
            if isinstance(f, models.ForeignKey):
                fields.append(f"{f.name}_id")
            else:
                fields.append(f.name)
        # Ensure id first if present
        if 'id' in fields:
            fields = ['id'] + [x for x in fields if x != 'id']
        return fields

    def _fix_arabic_encoding(self, text):
        """إصلاح ترميز النص العربي بطرق متعددة"""
        if not text:
            return ''
        
        # تحويل إلى string إذا لم يكن string
        if not isinstance(text, str):
            try:
                text = str(text)
            except:
                return ''
        
        # إذا كان النص يحتوي على رموز مثل Ø§Ø³Ø§، فهذا يعني أنه UTF-8 مخزن كـ Latin-1
        if 'Ø' in text or '§' in text or '³' in text or 'Ù' in text:
            try:
                # الطريقة 1: تحويل من Latin-1 إلى bytes ثم decode كـ UTF-8
                fixed = text.encode('latin-1').decode('utf-8')
                # التحقق من أن الإصلاح نجح (لا يجب أن يحتوي على رموز غريبة)
                if 'Ø' not in fixed and '§' not in fixed:
                    return fixed
            except (UnicodeEncodeError, UnicodeDecodeError):
                pass
            
            try:
                # الطريقة 2: استخدام errors='replace' أو 'ignore'
                fixed = text.encode('latin-1', errors='ignore').decode('utf-8', errors='ignore')
                if 'Ø' not in fixed and '§' not in fixed:
                    return fixed
            except:
                pass
        
        # إذا كان النص bytes، decodeه كـ UTF-8
        if isinstance(text, bytes):
            try:
                return text.decode('utf-8')
            except UnicodeDecodeError:
                try:
                    return text.decode('utf-8', errors='ignore')
                except:
                    return text.decode('latin-1', errors='ignore')
        
        return text

    def _get_value_for_export(self, obj, header):
        # If header is <field>_id, export related id
        if header.endswith('_id'):
            base = header[:-3]
            rel_id = getattr(obj, f"{base}_id", '') or ''
            # Ensure UUIDs are exported as strings
            if isinstance(rel_id, uuid.UUID):
                return str(rel_id)
            return rel_id
        val = getattr(obj, header, '')
        if isinstance(val, (datetime,)):
            return val.isoformat()
        # Ensure UUIDs are exported as strings
        if isinstance(val, uuid.UUID):
            return str(val)
        # Convert file/image fields to a simple string (file name)
        if isinstance(val, FieldFile):
            return self._fix_arabic_encoding(val.name) if val.name else ''
        # Fallback for any object with a file-like name/url attribute
        if hasattr(val, 'name') or hasattr(val, 'url'):
            try:
                name = getattr(val, 'name', None) or getattr(val, 'url', None)
                return self._fix_arabic_encoding(name) if name else ''
            except Exception:
                return ''
        # Apply Arabic encoding fix for string values
        if isinstance(val, str):
            return self._fix_arabic_encoding(val)
        return val if val is not None else ''

    def _parse_value(self, field_name, val):
        # Map back types using model field
        f = None
        for mf in self.model._meta.get_fields():
            if mf.name == field_name:
                f = mf
                break
        # If providing <field>_id, treat as integer/None
        if field_name.endswith('_id'):
            if val in (None, ''):
                return None
            try:
                return int(val)
            except Exception:
                return None
        if not f:
            return val
        try:
            if isinstance(f, (models.IntegerField, models.BigIntegerField, models.AutoField)):
                return int(val) if val not in (None, '') else None
            if isinstance(f, (models.FloatField, models.DecimalField)):
                return float(val) if val not in (None, '') else None
            if isinstance(f, models.BooleanField):
                if isinstance(val, bool):
                    return val
                if str(val).strip() in ('1', 'true', 'True', 'yes', 'نعم'):
                    return True
                if str(val).strip() in ('0', 'false', 'False', 'no', 'لا'):
                    return False
                return None
            if isinstance(f, (models.DateField, models.DateTimeField)):
                if not val:
                    return None
                # Try ISO formats
                for fmt in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S'):
                    try:
                        dt = datetime.strptime(str(val), fmt)
                        return dt if isinstance(f, models.DateTimeField) else dt.date()
                    except Exception:
                        continue
                return None
        except Exception:
            return None
        return val

    # ----- Views -----
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        opts = self.model._meta
        base = f"{opts.app_label}_{opts.model_name}"
        extra_context.update({
            'import_export_enabled': True,
            'template_url': reverse(f'admin:{base}_import_excel_template'),
            'import_excel_url': reverse(f'admin:{base}_import_excel'),
            'export_excel_url': reverse(f'admin:{base}_export_excel'),
        })
        return super().changelist_view(request, extra_context=extra_context)

    def download_excel_template(self, request):
        headers = self._model_headers()
        if Workbook:
            wb = Workbook()
            ws = wb.active
            ws.title = self.model._meta.verbose_name_plural.title()
            ws.append(headers)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{self.model._meta.model_name}_import_template.xlsx"'
            return response
        else:
            csv_content = ','.join(headers) + '\n'
            return HttpResponse(csv_content, content_type='text/csv')

    def export_excel_view(self, request):
        headers = self._model_headers()
        queryset = self.get_queryset(request)
        if Workbook:
            wb = Workbook()
            ws = wb.active
            ws.title = self.model._meta.verbose_name_plural.title()
            ws.append(headers)
            for obj in queryset:
                ws.append([self._get_value_for_export(obj, h) for h in headers])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{self.model._meta.model_name}_export.xlsx"'
            return response
        else:
            import codecs
            # CSV مع UTF-8 BOM لمساعدة Excel على قراءة النصوص العربية
            csv_content = codecs.BOM_UTF8.decode('utf-8')
            csv_content += ','.join(headers) + '\n'
            for obj in queryset:
                row = [str(self._get_value_for_export(obj, h)) for h in headers]
                csv_content += ','.join(row) + '\n'
            return HttpResponse(csv_content, content_type='text/csv; charset=utf-8')

    def import_excel_view(self, request):
        base_ctx = self.admin_site.each_context(request)
        context = {
            **base_ctx,
            'title': f'استيراد {self.model._meta.verbose_name_plural} من ملف إكسل',
            'opts': self.model._meta,
            'template_url': reverse(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import_excel_template'),
            'changelist_url': reverse(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist'),
        }

        if request.method == 'POST':
            file = request.FILES.get('excel_file')
            if not file:
                messages.error(request, 'يرجى اختيار ملف إكسل.')
                return redirect(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import_excel')
            if not load_workbook:
                messages.error(request, 'حزمة openpyxl غير متاحة. يرجى تثبيتها أو استخدام قالب CSV.')
                return redirect(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import_excel')

            try:
                wb = load_workbook(filename=file, data_only=True)
                ws = wb.active
            except Exception as e:
                messages.error(request, f'فشل قراءة الملف: {e}')
                return redirect(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import_excel')

            # Header row
            header_row = None
            for row in ws.iter_rows(values_only=True):
                header_row = [str(x).strip() if x is not None else '' for x in row]
                break
            if not header_row:
                messages.error(request, 'الملف لا يحتوي على صف رؤوس.')
                return redirect(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import_excel')

            headers = [h for h in header_row if h]
            created_count = 0
            updated_count = 0

            # Process rows
            for row in ws.iter_rows(values_only=True, min_row=2):
                data = {}
                for idx, h in enumerate(headers):
                    val = row[idx] if idx < len(row) else None
                    # Map <field>_id to ForeignKey id
                    data[h] = self._parse_value(h, val)

                # Determine identity by id if provided
                obj = None
                obj_id = data.get('id', None)
                if obj_id:
                    try:
                        obj = self.model.objects.get(id=obj_id)
                    except self.model.DoesNotExist:
                        obj = None

                # Prepare assignment dictionary for model fields
                assign = {}
                for k, v in data.items():
                    # If k endswith _id, assign to that related id field
                    if k.endswith('_id'):
                        assign[k] = v
                        continue
                    # Ensure the field exists
                    if any(f.name == k for f in self.model._meta.get_fields()):
                        assign[k] = v

                if obj:
                    for k, v in assign.items():
                        try:
                            setattr(obj, k, v)
                        except Exception:
                            pass
                    try:
                        obj.save()
                        updated_count += 1
                    except Exception:
                        pass
                else:
                    try:
                        obj = self.model.objects.create(**assign)
                        created_count += 1
                    except Exception:
                        pass

            messages.success(request, f'تم الاستيراد بنجاح. المُحدّث: {updated_count}، المُضاف: {created_count}.')
            return redirect(f'admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist')

        return TemplateResponse(request, 'admin/import_export/upload.html', context)